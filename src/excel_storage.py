import os
import datetime
from openpyxl import Workbook, load_workbook
from .utils import setup_logger, retry_operation

logger = setup_logger("excel_storage")

@retry_operation(max_retries=3, delay=1, logger=logger)
def save_to_local_excel(title, post_content):
    """Guarda en Excel local en .tmp/."""
    folder = ".tmp"
    filename = "posts_linkedin.xlsx"
    filepath = os.path.join(folder, filename)

    if not os.path.exists(folder): os.makedirs(folder)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.isfile(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Posts LinkedIn"
        ws.append(["Fecha", "Título", "Post"])
    else:
        wb = load_workbook(filepath)
        ws = wb.active

    ws.append([now, title, post_content])
    
    try:
        wb.save(filepath)
    except PermissionError:
        raise PermissionError(f"Cerrá el archivo '{filename}' antes de continuar.")

    return os.path.abspath(filepath)
